"""XLSX parser with chart analysis and multi-modal support"""

import io
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, ScatterChart
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

from core.parsers.interfaces import (
    BaseParser,
    Document,
    DocumentMetadata,
    DocumentType,
    ParseError,
    Segment,
    SegmentType,
    TextSubtype,
    VisualSubtype,
    TableSubtype,
    MetadataSubtype,
    VisualElement,
    VisualElementType,
)

logger = logging.getLogger(__name__)


class XLSXParser(BaseParser):
    """
    XLSX parser with chart analysis and multi-modal support
    
    Features:
    - Sheet-based text extraction
    - Chart extraction and analysis
    - Image extraction from spreadsheets
    - Multi-modal content with VLM descriptions
    - Cell-range based segmentation
    """
    
    def __init__(self, config: Optional[Dict[str, Any]] = None, enable_vlm: bool = True):
        """Initialize XLSX parser"""
        super().__init__(config, enable_vlm)
        self.supported_extensions = {".xlsx", ".xls"}
        
        # XLSX-specific configuration
        self.config.setdefault("extract_charts", True)
        self.config.setdefault("extract_images", True)
        self.config.setdefault("include_formulas", True)
        self.config.setdefault("max_rows_per_segment", 1)  # Default to one row per segment for better granularity
        self.config.setdefault("skip_empty_cells", True)
        self.config.setdefault("include_sheet_names", True)
        self.config.setdefault("segment_by_row", True)  # New option for row-level segmentation
        
        logger.info(f"Initialized XLSX parser with VLM: {enable_vlm}")
    
    def can_parse(self, file_path: Path) -> bool:
        """Check if file is an XLSX spreadsheet"""
        return file_path.suffix.lower() in self.supported_extensions
    
    async def parse(self, file_path: Path) -> Document:
        """
        Parse XLSX document with chart and image extraction
        
        Args:
            file_path: Path to XLSX file
            
        Returns:
            Document with text segments and visual elements
            
        Raises:
            ParseError: If parsing fails
        """
        self.validate_file(file_path)
        
        try:
            logger.info(f"Starting XLSX parsing: {file_path.name}")
            
            # Load workbook
            workbook = openpyxl.load_workbook(str(file_path), data_only=True)
            
            # Extract metadata
            metadata = self._extract_xlsx_metadata(file_path, workbook)
            
            # Extract visual elements
            visual_elements = []
            if self.config.get("extract_charts", True) or self.config.get("extract_images", True):
                visual_elements = self._extract_visual_elements(file_path, workbook)
            
            # Analyze visual elements with VLM
            if visual_elements and self.enable_vlm:
                document_context = {
                    "document_title": metadata.title,
                    "document_type": "XLSX",
                    "sheet_count": len(workbook.sheetnames)
                }
                visual_elements = await self.analyze_visual_elements(visual_elements, document_context)
            
            # Create segments from spreadsheet content
            segments = self._create_segments(workbook, visual_elements)
            
            # Create visual segments
            visual_segments = self._create_visual_segments(visual_elements)
            
            # Merge all segments
            all_segments = segments + visual_segments
            
            # Re-index segments
            for i, segment in enumerate(all_segments):
                segment.segment_index = i
            
            # Build full document content
            full_content = self._build_document_content(all_segments, visual_elements)
            
            # Create document
            document = self.create_document(
                file_path,
                segments=all_segments,
                metadata=metadata,
                visual_elements=visual_elements,
                content=full_content,
                raw_data=workbook
            )
            
            logger.info(f"XLSX parsing completed: {document.total_segments} segments, "
                       f"{document.total_visual_elements} visual elements")
            
            return document
            
        except Exception as e:
            logger.error(f"XLSX parsing failed: {e}", exc_info=True)
            raise ParseError(f"Failed to parse XLSX {file_path.name}: {str(e)}")
    
    def _extract_xlsx_metadata(self, file_path: Path, workbook: openpyxl.Workbook) -> DocumentMetadata:
        """Extract metadata from XLSX workbook"""
        try:
            # Get base metadata
            metadata = self.extract_metadata(file_path)
            metadata.document_type = DocumentType.XLSX
            
            # Extract workbook properties
            props = workbook.properties
            if props:
                metadata.title = props.title or metadata.title
                metadata.author = props.creator
                metadata.created_date = props.created
                metadata.modified_date = props.modified
            
            # Count sheets as pages
            metadata.page_count = len(workbook.sheetnames)
            
            # Add custom metadata
            metadata.custom_metadata.update({
                "sheets": workbook.sheetnames,
                "sheet_count": len(workbook.sheetnames),
                "has_charts": self._has_charts(workbook),
                "has_images": self._has_images(workbook),
                "format": "XLSX"
            })
            
            return metadata
            
        except Exception as e:
            logger.warning(f"Failed to extract XLSX metadata: {e}")
            metadata = self.extract_metadata(file_path)
            metadata.document_type = DocumentType.XLSX
            return metadata
    
    def _has_charts(self, workbook: openpyxl.Workbook) -> bool:
        """Check if workbook contains charts"""
        try:
            for sheet in workbook.worksheets:
                if hasattr(sheet, '_charts') and sheet._charts:
                    return True
            return False
        except Exception:
            return False
    
    def _has_images(self, workbook: openpyxl.Workbook) -> bool:
        """Check if workbook contains images"""
        try:
            for sheet in workbook.worksheets:
                if hasattr(sheet, '_images') and sheet._images:
                    return True
            return False
        except Exception:
            return False
    
    def _extract_visual_elements(self, file_path: Path, workbook: openpyxl.Workbook) -> List[VisualElement]:
        """Extract visual elements from XLSX workbook"""
        visual_elements = []
        
        try:
            # Extract from each sheet
            for sheet_idx, sheet in enumerate(workbook.worksheets):
                sheet_name = sheet.title
                
                # Extract charts
                if self.config.get("extract_charts", True):
                    visual_elements.extend(self._extract_charts_from_sheet(sheet, sheet_name, sheet_idx))
                
                # Extract images
                if self.config.get("extract_images", True):
                    visual_elements.extend(self._extract_images_from_sheet(sheet, sheet_name, sheet_idx))
            
            logger.info(f"Extracted {len(visual_elements)} visual elements from XLSX")
            return visual_elements
            
        except Exception as e:
            logger.error(f"Visual element extraction failed: {e}")
            return []
    
    def _extract_charts_from_sheet(self, sheet, sheet_name: str, sheet_idx: int) -> List[VisualElement]:
        """Extract charts from a worksheet"""
        visual_elements = []
        
        try:
            if not hasattr(sheet, '_charts') or not sheet._charts:
                return visual_elements
            
            for chart_idx, chart in enumerate(sheet._charts):
                try:
                    # Create chart visualization data
                    chart_data = self._analyze_chart(chart)
                    
                    # Create a simple representation of the chart
                    chart_text = self._chart_to_text(chart, chart_data)
                    chart_bytes = chart_text.encode('utf-8')
                    
                    # Determine chart type
                    chart_type = self._determine_chart_type(chart)
                    
                    visual_element = VisualElement(
                        element_type=chart_type,
                        source_format=DocumentType.XLSX,
                        content_hash=VisualElement.create_hash(chart_bytes),
                        raw_data=chart_bytes,  # Store as text representation
                        page_or_slide=sheet_idx + 1,
                        segment_reference=f"sheet_{sheet_name}",
                        file_extension="txt",  # Text representation
                        extracted_data=chart_data,
                        analysis_metadata={
                            "sheet_name": sheet_name,
                            "chart_index": chart_idx,
                            "chart_type": type(chart).__name__,
                            "extraction_method": "openpyxl_chart"
                        }
                    )
                    
                    visual_elements.append(visual_element)
                    
                except Exception as e:
                    logger.warning(f"Failed to extract chart {chart_idx} from sheet {sheet_name}: {e}")
                    continue
            
        except Exception as e:
            logger.error(f"Chart extraction failed for sheet {sheet_name}: {e}")
        
        return visual_elements
    
    def _extract_images_from_sheet(self, sheet, sheet_name: str, sheet_idx: int) -> List[VisualElement]:
        """Extract images from a worksheet"""
        visual_elements = []
        
        try:
            if not hasattr(sheet, '_images') or not sheet._images:
                return visual_elements
            
            for img_idx, img in enumerate(sheet._images):
                try:
                    # Get image data
                    image_data = self._get_image_data(img)
                    if image_data and len(image_data) > 1000:  # Minimum size check
                        
                        # Determine image type
                        img_type = self._determine_image_type(image_data)
                        
                        visual_element = VisualElement(
                            element_type=img_type,
                            source_format=DocumentType.XLSX,
                            content_hash=VisualElement.create_hash(image_data),
                            raw_data=image_data,
                            page_or_slide=sheet_idx + 1,
                            segment_reference=f"sheet_{sheet_name}",
                            file_extension=self._get_image_extension(image_data),
                            analysis_metadata={
                                "sheet_name": sheet_name,
                                "image_index": img_idx,
                                "extraction_method": "openpyxl_image"
                            }
                        )
                        
                        visual_elements.append(visual_element)
                        
                except Exception as e:
                    logger.warning(f"Failed to extract image {img_idx} from sheet {sheet_name}: {e}")
                    continue
            
        except Exception as e:
            logger.error(f"Image extraction failed for sheet {sheet_name}: {e}")
        
        return visual_elements
    
    def _analyze_chart(self, chart) -> Dict[str, Any]:
        """Analyze chart and extract data"""
        try:
            chart_data = {
                "chart_type": type(chart).__name__,
                "title": getattr(chart, 'title', None),
                "series_count": len(getattr(chart, 'series', [])),
                "categories": [],
                "values": [],
                "data_points": []
            }
            
            # Extract chart title
            if hasattr(chart, 'title') and chart.title:
                chart_data["title"] = str(chart.title)
            
            # Extract series data
            if hasattr(chart, 'series'):
                for series in chart.series:
                    try:
                        series_data = {
                            "title": getattr(series, 'title', None),
                            "values": []
                        }
                        
                        # Extract values if available
                        if hasattr(series, 'values') and series.values:
                            series_data["values"] = self._extract_series_values(series.values)
                        
                        chart_data["data_points"].append(series_data)
                        
                    except Exception as e:
                        logger.debug(f"Failed to extract series data: {e}")
                        continue
            
            return chart_data
            
        except Exception as e:
            logger.debug(f"Chart analysis failed: {e}")
            return {"chart_type": "unknown", "error": str(e)}
    
    def _extract_series_values(self, values_ref) -> List[Any]:
        """Extract values from chart series reference"""
        try:
            if hasattr(values_ref, 'cache') and values_ref.cache:
                return [str(pt.v) for pt in values_ref.cache if hasattr(pt, 'v')]
            return []
        except Exception:
            return []
    
    def _determine_chart_type(self, chart) -> VisualElementType:
        """Determine visual element type based on chart type"""
        chart_type_map = {
            'BarChart': VisualElementType.CHART,
            'LineChart': VisualElementType.GRAPH,
            'PieChart': VisualElementType.CHART,
            'ScatterChart': VisualElementType.GRAPH,
            'AreaChart': VisualElementType.CHART,
        }
        
        chart_class = type(chart).__name__
        return chart_type_map.get(chart_class, VisualElementType.CHART)
    
    def _chart_to_text(self, chart, chart_data: Dict[str, Any]) -> str:
        """Convert chart to text representation"""
        try:
            text_parts = []
            
            # Add title
            if chart_data.get("title"):
                text_parts.append(f"Chart Title: {chart_data['title']}")
            
            # Add chart type
            text_parts.append(f"Chart Type: {chart_data.get('chart_type', 'Unknown')}")
            
            # Add series information
            if chart_data.get("data_points"):
                text_parts.append("Data Series:")
                for i, series in enumerate(chart_data["data_points"]):
                    series_title = series.get("title", f"Series {i+1}")
                    values = series.get("values", [])
                    if values:
                        text_parts.append(f"  {series_title}: {', '.join(values[:10])}")  # First 10 values
                    else:
                        text_parts.append(f"  {series_title}: No data")
            
            return "\n".join(text_parts)
            
        except Exception as e:
            logger.debug(f"Chart to text conversion failed: {e}")
            return f"Chart: {chart_data.get('chart_type', 'Unknown')}"
    
    def _get_image_data(self, img: XLImage) -> Optional[bytes]:
        """Get image data from Excel image object"""
        try:
            if hasattr(img, '_data') and img._data:
                return img._data()
            elif hasattr(img, 'ref') and img.ref:
                # Try to get data from reference
                return None  # Would need more complex extraction
            return None
        except Exception as e:
            logger.debug(f"Failed to get image data: {e}")
            return None
    
    def _determine_image_type(self, image_data: bytes) -> VisualElementType:
        """Determine visual element type for image"""
        try:
            img = Image.open(io.BytesIO(image_data))
            width, height = img.size
            
            # Heuristics for image classification
            aspect_ratio = width / height
            
            if aspect_ratio > 2.0:
                return VisualElementType.CHART
            elif 0.8 <= aspect_ratio <= 1.2:
                return VisualElementType.DIAGRAM
            else:
                return VisualElementType.IMAGE
                
        except Exception:
            return VisualElementType.UNKNOWN_VISUAL
    
    def _get_image_extension(self, image_data: bytes) -> str:
        """Get file extension for image"""
        try:
            img = Image.open(io.BytesIO(image_data))
            format_map = {
                'JPEG': 'jpg',
                'PNG': 'png',
                'BMP': 'bmp',
                'GIF': 'gif',
                'TIFF': 'tiff'
            }
            return format_map.get(img.format, 'png')
        except Exception:
            return 'png'
    
    def _create_segments(self, workbook: openpyxl.Workbook, visual_elements: List[VisualElement]) -> List[Segment]:
        """Create segments from spreadsheet content"""
        segments = []
        
        try:
            segment_idx = 0
            
            for sheet_idx, sheet in enumerate(workbook.worksheets):
                sheet_name = sheet.title
                
                # Create sheet header segment
                if self.config.get("include_sheet_names", True):
                    sheet_segment = Segment(
                        content=f"Sheet: {sheet_name}",
                        segment_index=segment_idx,
                        segment_type=SegmentType.METADATA,
                        segment_subtype=MetadataSubtype.SHEET.value,
                        page_number=sheet_idx + 1,
                        metadata={
                            "sheet_name": sheet_name,
                            "sheet_index": sheet_idx
                        }
                    )
                    segments.append(sheet_segment)
                    segment_idx += 1
                
                # Process sheet content
                sheet_segments = self._process_sheet_content(sheet, sheet_name, sheet_idx, segment_idx, visual_elements)
                segments.extend(sheet_segments)
                segment_idx += len(sheet_segments)
            
            logger.info(f"Created {len(segments)} segments from XLSX")
            return segments
            
        except Exception as e:
            logger.error(f"Segment creation failed: {e}")
            return []
    
    def _process_sheet_content(self, sheet, sheet_name: str, sheet_idx: int, start_idx: int, visual_elements: List[VisualElement]) -> List[Segment]:
        """Process content from a single sheet"""
        segments = []
        
        try:
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            if max_row == 0 or max_col == 0:
                # Empty sheet
                return segments
            
            # Find visual references for this sheet
            sheet_visual_refs = []
            for ve in visual_elements:
                if ve.segment_reference == f"sheet_{sheet_name}":
                    sheet_visual_refs.append(ve.content_hash)
            
            # Check if we should segment by row
            segment_by_row = self.config.get("segment_by_row", True)
            max_rows_per_segment = self.config.get("max_rows_per_segment", 1)
            
            # Extract header row if present (first non-empty row)
            header_row = None
            header_row_idx = 0
            for row_idx in range(1, min(10, max_row + 1)):  # Check first 10 rows for header
                row_data = []
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col)
                    if cell.value is not None:
                        row_data.append(str(cell.value))
                if len(row_data) > 1:  # Found a row with multiple values
                    header_row = row_data
                    header_row_idx = row_idx
                    break
            
            # Process rows
            for row_idx in range(1, max_row + 1):
                # For row-level segmentation
                if segment_by_row and max_rows_per_segment == 1:
                    row_content = self._extract_single_row_content(sheet, row_idx, max_col)
                    
                    if row_content and row_content.strip():  # Only create segment if row has content
                        segment = Segment(
                            content=row_content,
                            segment_index=start_idx + len(segments),
                            segment_type=SegmentType.TABLE,
                            segment_subtype=TableSubtype.DATA.value if row_idx != header_row_idx else TableSubtype.HEADER.value,
                            page_number=sheet_idx + 1,
                            visual_references=sheet_visual_refs if row_idx == 1 else [],  # Only attach visual refs to first row
                            metadata={
                                "sheet_name": sheet_name,
                                "row_number": row_idx,
                                "column_count": max_col,
                                "is_header": row_idx == header_row_idx,
                                "has_charts": any(ve.element_type in [VisualElementType.CHART, VisualElementType.GRAPH] for ve in visual_elements if ve.segment_reference == f"sheet_{sheet_name}"),
                                "has_images": any(ve.element_type == VisualElementType.IMAGE for ve in visual_elements if ve.segment_reference == f"sheet_{sheet_name}")
                            }
                        )
                        segments.append(segment)
                else:
                    # Process in chunks for larger segments
                    if (row_idx - 1) % max_rows_per_segment == 0:
                        row_start = row_idx
                        row_end = min(row_start + max_rows_per_segment - 1, max_row)
                        
                        # Extract content from row range
                        range_content = self._extract_range_content(sheet, row_start, row_end, max_col)
                        
                        if range_content:
                            segment = Segment(
                                content=range_content,
                                segment_index=start_idx + len(segments),
                                segment_type=SegmentType.TABLE,
                                segment_subtype=TableSubtype.DATA.value,
                                page_number=sheet_idx + 1,
                                visual_references=sheet_visual_refs if row_start == 1 else [],
                                metadata={
                                    "sheet_name": sheet_name,
                                    "row_start": row_start,
                                    "row_end": row_end,
                                    "column_count": max_col,
                                    "has_charts": any(ve.element_type in [VisualElementType.CHART, VisualElementType.GRAPH] for ve in visual_elements if ve.segment_reference == f"sheet_{sheet_name}"),
                                    "has_images": any(ve.element_type == VisualElementType.IMAGE for ve in visual_elements if ve.segment_reference == f"sheet_{sheet_name}")
                                }
                            )
                            segments.append(segment)
            
        except Exception as e:
            logger.error(f"Sheet processing failed for {sheet_name}: {e}")
        
        return segments
    
    def _extract_single_row_content(self, sheet, row_idx: int, max_col: int) -> str:
        """Extract content from a single row"""
        try:
            row_data = []
            
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col)
                cell_value = cell.value
                
                if cell_value is not None:
                    # Include formulas if configured
                    if self.config.get("include_formulas", True) and str(cell_value).startswith('='):
                        row_data.append(f"[Formula: {cell_value}]")
                    else:
                        row_data.append(str(cell_value))
                elif not self.config.get("skip_empty_cells", True):
                    row_data.append("")
            
            # Return row content if it has any data
            if row_data and any(cell.strip() for cell in row_data if cell):
                return " | ".join(row_data)
            
            return ""
            
        except Exception as e:
            logger.debug(f"Single row extraction failed: {e}")
            return ""
    
    def _extract_range_content(self, sheet, row_start: int, row_end: int, max_col: int) -> str:
        """Extract content from a range of cells"""
        try:
            content_lines = []
            
            for row in range(row_start, row_end + 1):
                row_data = []
                
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_value = cell.value
                    
                    if cell_value is not None:
                        # Include formulas if configured
                        if self.config.get("include_formulas", True) and str(cell_value).startswith('='):
                            row_data.append(f"[Formula: {cell_value}]")
                        else:
                            row_data.append(str(cell_value))
                    elif not self.config.get("skip_empty_cells", True):
                        row_data.append("")
                
                # Add row if it has content
                if row_data and any(cell.strip() for cell in row_data if cell):
                    content_lines.append(" | ".join(row_data))
            
            return "\n".join(content_lines)
            
        except Exception as e:
            logger.debug(f"Range extraction failed: {e}")
            return ""
    
    def _build_document_content(self, segments: List[Segment], visual_elements: List[VisualElement]) -> str:
        """Build full document content"""
        content_parts = []
        
        # Add text content
        for segment in segments:
            if hasattr(segment.segment_type, 'value') and segment.segment_type == SegmentType.METADATA and segment.segment_subtype == MetadataSubtype.SHEET.value:
                content_parts.append(f"## {segment.content}")
            else:
                content_parts.append(segment.content)
        
        # Add visual element descriptions
        for visual in visual_elements:
            if visual.vlm_description:
                visual_desc = f"\n[VISUAL ELEMENT - {visual.element_type.value.upper()}]"
                if visual.segment_reference:
                    visual_desc += f" (in {visual.segment_reference})"
                visual_desc += f": {visual.vlm_description}"
                content_parts.append(visual_desc)
        
        return "\n\n".join(content_parts)
    
    def _create_visual_segments(self, visual_elements: List[VisualElement]) -> List[Segment]:
        """Create segments from visual elements"""
        visual_segments = []
        
        for visual_elem in visual_elements:
            # Determine visual subtype
            if visual_elem.element_type == VisualElementType.CHART:
                subtype = VisualSubtype.CHART.value
            elif visual_elem.element_type == VisualElementType.GRAPH:
                subtype = VisualSubtype.CHART.value  # Map GRAPH to CHART
            elif visual_elem.element_type == VisualElementType.IMAGE:
                subtype = VisualSubtype.IMAGE.value
            elif visual_elem.element_type == VisualElementType.DIAGRAM:
                subtype = VisualSubtype.DIAGRAM.value
            else:
                subtype = VisualSubtype.CHART.value  # Default for Excel visuals
            
            # Create placeholder content
            content = f"[{visual_elem.element_type.value.upper()}: Placeholder]"
            
            # Create visual segment
            segment = Segment(
                content=content,
                page_number=visual_elem.page_or_slide,  # sheet index as page
                segment_type=SegmentType.VISUAL,
                segment_subtype=subtype,
                visual_references=[visual_elem.content_hash],
                metadata={
                    "visual_type": visual_elem.element_type.value,
                    "sheet_name": visual_elem.analysis_metadata.get("sheet_name", ""),
                    "extraction_method": visual_elem.analysis_metadata.get("extraction_method", ""),
                    "source": "xlsx_visual_extraction"
                }
            )
            
            visual_segments.append(segment)
        
        return visual_segments